from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, delete
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime

from app.api.deps import get_session, get_current_user, require_feature
from app.models.user import User
from app.models.tickets import Ticket, TicketCategory, TicketInboxShare, ShareStatus
from app.schemas.tickets import (
    TicketRead, TicketCreate, TicketUpdate,
    TicketCategoryRead, TicketCategoryCreate, TicketCategoryUpdate,
    TicketInboxInviteCreate, TicketInboxShareRead
)
from app.core.features import Feature
from app.models.phone import Phone

router = APIRouter(dependencies=[Depends(require_feature(Feature.tickets))])

# ── Helper ──
def get_accessible_owner_ids(session: Session, user_id: int) -> list[int]:
    """Returns the user's ID plus any owner IDs who have shared their inbox with this user."""
    shared_inboxes = session.exec(
        select(TicketInboxShare)
        .where(TicketInboxShare.shared_with_id == user_id)
        .where(TicketInboxShare.status == ShareStatus.ACCEPTED)
    ).all()
    return [user_id] + [s.owner_id for s in shared_inboxes]

# ── Categories ──
@router.post("/categories", response_model=TicketCategoryRead)
def create_category(
    data: TicketCategoryCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    category = TicketCategory(**data.model_dump(), user_id=current_user.id)
    session.add(category)
    session.commit()
    session.refresh(category)
    return category

@router.get("/categories", response_model=list[TicketCategoryRead])
def get_categories(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    owner_ids = get_accessible_owner_ids(session, current_user.id)
    categories = session.exec(
        select(TicketCategory).where(TicketCategory.user_id.in_(owner_ids))
    ).all()
    return categories

@router.put("/categories/{cat_id}", response_model=TicketCategoryRead)
def update_category(
    cat_id: int,
    data: TicketCategoryUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    category = session.get(TicketCategory, cat_id)
    if not category or category.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Category not found or unauthorized")
        
    if data.name is not None:
        category.name = data.name
    if data.color is not None:
        category.color = data.color
        
    session.add(category)
    session.commit()
    session.refresh(category)
    return category

@router.delete("/categories/{cat_id}", status_code=204)
def delete_category(
    cat_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    category = session.get(TicketCategory, cat_id)
    if not category or category.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Category not found or unauthorized")
    session.delete(category)
    session.commit()

# ── Internal Service Logic ──
def create_ticket_internal(session: Session, user_id: int, sender_number: str, body: str, category_id: int = None, session_id: str = None) -> Ticket:
    """Internal method for flows/bots to create tickets. Does not require HTTP request context."""
    if category_id:
        cat = session.get(TicketCategory, category_id)
        if not cat or cat.user_id != user_id:
            category_id = None # Ignore invalid category gracefully

    ticket = Ticket(
        user_id=user_id,
        sender_number=sender_number,
        session_id=session_id,
        body=body,
        category_id=category_id
    )
    session.add(ticket)
    session.commit()
    
    # Reload with relations if needed
    return session.exec(
        select(Ticket)
        .where(Ticket.id == ticket.id)
        .options(selectinload(Ticket.category), selectinload(Ticket.user))
    ).first()

@router.get("/export")
def export_tickets_xlsx(
    mode: str = Query("custom", pattern="^(new|custom)$"),
    start_date: datetime = Query(None),
    end_date: datetime = Query(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    owner_ids = get_accessible_owner_ids(session, current_user.id)
    query = select(Ticket).where(Ticket.user_id.in_(owner_ids)).options(selectinload(Ticket.category), selectinload(Ticket.user))
    
    if mode == "new" and current_user.last_ticket_download_at:
        query = query.where(Ticket.created_at > current_user.last_ticket_download_at)
    
    if start_date:
        query = query.where(Ticket.created_at >= start_date)
    if end_date:
        query = query.where(Ticket.created_at <= end_date)
        
    tickets = session.exec(query.order_by(Ticket.created_at.desc())).all()
    
    # ── Optimized Phone Lookup ──
    # Only fetch phones that appear in the results
    session_ids = {t.session_id for t in tickets if t.session_id}
    phone_map = {}
    if session_ids:
        phones = session.exec(select(Phone).where(Phone.session_id.in_(list(session_ids)))).all()
        phone_map = {p.session_id: p.name for p in phones}

    # ── Build Workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets Export"
    
    # Common Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Violet 600
    header_font = Font(bold=True, size=14, color="FFFFFF")
    data_font = Font(size=12)
    thin_border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Header
    headers = ["Date", "Phone", "Sender Number", "Body", "Category", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    for index, t in enumerate(tickets, start=2):
        phone_name = phone_map.get(t.session_id, "N/A") if t.session_id else "N/A"
        
        # Build Data Row
        row_data = [
            t.created_at.strftime("%Y-%m-%d %H:%M"),
            phone_name,
            t.sender_number,
            t.body,
            t.category.name if t.category else "None",
            "Open" if t.is_open else "Closed"
        ]
        ws.append(row_data)
        
        # Apply Styles to the new Row
        for col_idx, cell in enumerate(ws[index], start=1):
            cell.font = data_font
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 2, 3, 5, 6]:
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align # Body
            
            # Category Color Fill
            if col_idx == 5 and t.category and t.category.color:
                # Remove '#' if present
                hex_color = t.category.color.lstrip('#')
                # Ensure it's valid HEX for openpyxl
                if len(hex_color) == 6:
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    # Determine text color (white for dark backgrounds)
                    cell.font = Font(size=12, color="FFFFFF", bold=True)
        
    # Auto-adjust column widths
    column_widths = {
        1: 20, # Date
        2: 20, # Phone
        3: 20, # Sender
        4: 60, # Body (wider + wrap)
        5: 15, # Category
        6: 12  # Status
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Update Last Download Date
    current_user.last_ticket_download_at = datetime.utcnow()
    session.add(current_user)
    session.commit()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"tickets_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename=\"{filename}\"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# ── Tickets ──
@router.get("/", response_model=list[TicketRead])
def get_tickets(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    owner_ids = get_accessible_owner_ids(session, current_user.id)
    
    tickets = session.exec(
        select(Ticket)
        .where(Ticket.user_id.in_(owner_ids))
        .options(selectinload(Ticket.category), selectinload(Ticket.user))
        .order_by(Ticket.is_open.desc(), Ticket.created_at.desc())
    ).all()

    # Resolve phone names
    results = []
    for t in tickets:
        read_data = TicketRead.model_validate(t)
        if t.session_id:
            phone = session.exec(select(Phone).where(Phone.session_id == t.session_id)).first()
            if phone:
                read_data.phone_name = phone.name
        results.append(read_data)
        
    return results

@router.put("/{ticket_id}", response_model=TicketRead)
def update_ticket(
    ticket_id: int,
    data: TicketUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    ticket = session.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
        
    owner_ids = get_accessible_owner_ids(session, current_user.id)
    if ticket.user_id not in owner_ids:
        raise HTTPException(status_code=403, detail="Not authorized to modify this ticket")
        
    if data.is_open is not None:
        ticket.is_open = data.is_open
    if data.category_id is not None:
        ticket.category_id = data.category_id
        
    session.add(ticket)
    session.commit()
    session.refresh(ticket)
    
    # Return as Read schema with phone_name resolved
    res = TicketRead.model_validate(ticket)
    if ticket.session_id:
        phone = session.exec(select(Phone).where(Phone.session_id == ticket.session_id)).first()
        if phone:
            res.phone_name = phone.name
            
    return res

@router.delete("/{ticket_id}", status_code=204)
def delete_ticket(
    ticket_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    ticket = session.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
        
    if ticket.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Only the owner can delete this ticket")
        
    session.delete(ticket)
    session.commit()

# ── Shares (Invitations) ──
@router.post("/share/invite", response_model=TicketInboxShareRead)
def send_invite(
    data: TicketInboxInviteCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    if data.email.lower() == current_user.email.lower():
        raise HTTPException(status_code=400, detail="Cannot invite yourself")
        
    target_user = session.exec(select(User).where(User.email == data.email)).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User with this email not found")
        
    existing = session.exec(
        select(TicketInboxShare)
        .where(TicketInboxShare.owner_id == current_user.id)
        .where(TicketInboxShare.shared_with_id == target_user.id)
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="Invite already exists for this user")
        
    share = TicketInboxShare(owner_id=current_user.id, shared_with_id=target_user.id)
    session.add(share)
    session.commit()
    
    return session.exec(
        select(TicketInboxShare)
        .where(TicketInboxShare.id == share.id)
        .options(selectinload(TicketInboxShare.shared_with), selectinload(TicketInboxShare.owner))
    ).first()

@router.get("/share/invites", response_model=list[TicketInboxShareRead])
def get_invites(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Get all received invites to manage other inboxes."""
    invites = session.exec(
        select(TicketInboxShare)
        .where(TicketInboxShare.shared_with_id == current_user.id)
        .options(selectinload(TicketInboxShare.shared_with), selectinload(TicketInboxShare.owner))
    ).all()
    return invites
    
@router.get("/share/sent", response_model=list[TicketInboxShareRead])
def get_sent_invites(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Get all outgoing invites we sent out to other users."""
    invites = session.exec(
        select(TicketInboxShare)
        .where(TicketInboxShare.owner_id == current_user.id)
        .options(selectinload(TicketInboxShare.shared_with), selectinload(TicketInboxShare.owner))
    ).all()
    return invites

@router.put("/share/invites/{share_id}/{action}", response_model=TicketInboxShareRead)
def manage_invite(
    share_id: int,
    action: str,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    if action not in ["accept", "reject"]:
        raise HTTPException(status_code=400, detail="Invalid action")
        
    share = session.get(TicketInboxShare, share_id)
    if not share or share.shared_with_id != current_user.id:
        raise HTTPException(status_code=404, detail="Invite not found")
        
    share.status = ShareStatus.ACCEPTED if action == "accept" else ShareStatus.REJECTED
    session.add(share)
    session.commit()
    
    return session.exec(
        select(TicketInboxShare)
        .where(TicketInboxShare.id == share.id)
        .options(selectinload(TicketInboxShare.shared_with), selectinload(TicketInboxShare.owner))
    ).first()

@router.delete("/share/{share_id}")
def revoke_share(
    share_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Owner can totally revoke a share they previously sent."""
    share = session.get(TicketInboxShare, share_id)
    if not share or share.owner_id != current_user.id:
        raise HTTPException(status_code=404, detail="Share not found or unauthorized")
        
    session.delete(share)
    session.commit()
    return {"ok": True}
