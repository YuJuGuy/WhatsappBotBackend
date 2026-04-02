from fastapi import APIRouter, Depends, HTTPException, status
from typing import List
from sqlmodel import Session, select, delete
from datetime import datetime, timedelta, timezone

from app.schemas.flow import FlowCreate, FlowRead, FlowListRead, FlowUpdate
from app.core.features import Feature
from app.api.deps import get_session, get_current_user, require_feature, user_has_feature
from app.models.user import User
from app.models.phone import Phone
from app.models.flow import Flow, FlowNode, FlowEdge, FlowRun, FlowRunStatus, FlowPhoneLink
from app.api.rate_limit import rate_limit_by_user
from app.api.outbox.routes import insert_outbox
from app.models.storage import StoredFile
from app.core.storage import USER_FILES_ROOT
import openpyxl

router = APIRouter(dependencies=[Depends(require_feature(Feature.flows))])

def _resolve_variables(text: str, run: FlowRun, current_message: str) -> str:
    if not text:
        return ""
    
    # Standard variables
    data = run.session_metadata or {}
    variables = {
        "message": current_message,
        "sender": run.contact_id.split("@")[0] if "@" in run.contact_id else run.contact_id,
        "now": datetime.now().strftime("%Y-%m-%d %H:%M"),
        **data
    }
    
    # Apply replacements
    for key, val in variables.items():
        placeholder = "{{" + key + "}}"
        if placeholder in text:
            text = text.replace(placeholder, str(val if val is not None else ""))
    
    return text

def _build_flow_response(flow: Flow, session: Session) -> dict:
    nodes = session.exec(select(FlowNode).where(FlowNode.flow_id == flow.id)).all()
    edges = session.exec(select(FlowEdge).where(FlowEdge.flow_id == flow.id)).all()
    phone_links = session.exec(select(FlowPhoneLink).where(FlowPhoneLink.flow_id == flow.id)).all()
    
    # Reverse map for frontend: DB int ID -> React string ID
    int_to_client = {n.id: n.client_node_id for n in nodes}
    
    return {
        "id": flow.id,
        "name": flow.name,
        "is_active": flow.is_active,
        "priority": flow.priority,
        "message_priority": flow.message_priority,
        "priority_over_autoreply": flow.priority_over_autoreply,
        "timeout_minutes": flow.timeout_minutes,
        "share_code": flow.share_code,
        "phone_ids": [l.phone_id for l in phone_links],
        "created_at": flow.created_at,
        "updated_at": flow.updated_at,
        "nodes": [
            {
                "node_id": n.client_node_id,
                "node_type": n.node_type,
                "node_data": n.node_data
            } for n in nodes
        ],
        "edges": [
            {
                "edge_id": e.client_edge_id,
                "source_node_id": int_to_client.get(e.from_node_id, str(e.from_node_id)),
                "target_node_id": int_to_client.get(e.to_node_id, str(e.to_node_id)),
                "edge_type": e.edge_type,
                "edge_data": e.edge_data
            } for e in edges
        ]
    }

def _validate_exactly_one_start_node(nodes):
    """Ensure exactly one START node exists in the payload, but allow empty shell creation."""
    if not nodes:
        return
    start_count = sum(1 for n in nodes if n.node_type == "start")
    if start_count != 1:
        raise HTTPException(status_code=400, detail=f"Flow must contain exactly 1 start node. Found {start_count}.")

def _validate_action_nodes(nodes, user: User):
    """Reject flows with action nodes the user doesn't have the feature for."""
    if not nodes:
        return
    has_ticket_node = any(n.node_type == "action_ticket" for n in nodes)
    if has_ticket_node and not user_has_feature(user, Feature.tickets):
        raise HTTPException(status_code=403, detail="You don't have access to the Tickets feature. Remove the 'Create Ticket' node.")


def _sync_phone_links(session: Session, flow_id: int, phone_ids: List[int]):
    session.exec(delete(FlowPhoneLink).where(FlowPhoneLink.flow_id == flow_id))
    session.flush()
    if phone_ids:
        new_links = [FlowPhoneLink(flow_id=flow_id, phone_id=pid) for pid in phone_ids]
        session.add_all(new_links)


def _sync_graph(session: Session, flow_id: int, nodes_in: list, edges_in: list):
    """Replace entire graph contents with new ones via highly-optimized bulk operations."""
    
    # 1. Bulk instant SQL deletes (Zero memory overhead)
    # Must delete edges before nodes due to Foreign Key referential integrity!
    session.exec(delete(FlowEdge).where(FlowEdge.flow_id == flow_id))
    session.exec(delete(FlowNode).where(FlowNode.flow_id == flow_id))

    # 2. Bulk inserts for Nodes
    new_nodes = [
        FlowNode(
            flow_id=flow_id,
            client_node_id=n.node_id,
            node_type=n.node_type,
            node_data=n.node_data.model_dump() if hasattr(n.node_data, "model_dump") else n.node_data
        ) for n in nodes_in
    ]
    session.add_all(new_nodes)
    session.flush() # Flush to populate auto-incrementing .id on all Node objects 
    
    # Map React string ID -> DB int ID
    node_id_map = {n.client_node_id: n.id for n in new_nodes}
        
    # 3. Bulk inserts for Edges, using the newly generated integer relationships
    new_edges = [
        FlowEdge(
            flow_id=flow_id,
            client_edge_id=e.edge_id,
            from_node_id=node_id_map[e.source_node_id],
            to_node_id=node_id_map[e.target_node_id],
            edge_type=e.edge_type,
            edge_data=e.edge_data
        ) for e in edges_in if e.source_node_id in node_id_map and e.target_node_id in node_id_map
    ]
    session.add_all(new_edges)


@router.post("/", response_model=FlowRead, status_code=status.HTTP_201_CREATED, dependencies=[Depends(rate_limit_by_user(20, 60, "flow-create"))])
def create_flow(flow_in: FlowCreate, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Create a new conversational flow and its graph nodes/edges."""
    _validate_exactly_one_start_node(flow_in.nodes)
    _validate_action_nodes(flow_in.nodes, current_user)
    
    flow = Flow(
        user_id=current_user.id,
        name=flow_in.name,
        is_active=flow_in.is_active,
        priority=flow_in.priority,
        message_priority=flow_in.message_priority,
        priority_over_autoreply=flow_in.priority_over_autoreply,
        timeout_minutes=flow_in.timeout_minutes
    )
    session.add(flow)
    session.commit()
    session.refresh(flow)
    
    _sync_phone_links(session, flow.id, flow_in.phone_ids)
    _sync_graph(session, flow.id, flow_in.nodes, flow_in.edges)
    session.commit()
    
    return _build_flow_response(flow, session)


@router.get("/", response_model=List[FlowListRead])
def get_flows(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Get heavily optimized list of user flows (excludes the giant node map)."""
    flows = session.exec(select(Flow).where(Flow.user_id == current_user.id).order_by(Flow.created_at.desc())).all()
    if not flows:
        return []
        
    flow_ids = [f.id for f in flows]
    all_links = session.exec(select(FlowPhoneLink).where(FlowPhoneLink.flow_id.in_(flow_ids))).all()
    phone_map = {fid: [] for fid in flow_ids}
    for link in all_links:
        phone_map[link.flow_id].append(link.phone_id)
        
    res = []
    for flow in flows:
        fd = flow.model_dump()
        fd["phone_ids"] = phone_map[flow.id]
        res.append(fd)
        
    return res


@router.get("/{flow_id}", response_model=FlowRead)
def get_flow(flow_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Get a specific flow with its full node/edge map."""
    flow = session.get(Flow, flow_id)
    if not flow or flow.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Flow not found")
    return _build_flow_response(flow, session)


@router.put("/{flow_id}", response_model=FlowRead, dependencies=[Depends(rate_limit_by_user(20, 60, "flow-update"))])
def update_flow(flow_id: int, flow_in: FlowUpdate, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Update a flow. If nodes/edges are provided, replaces the old graph completely."""
    flow = session.get(Flow, flow_id)
    if not flow or flow.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Flow not found")

    if flow_in.name is not None:
        flow.name = flow_in.name
    if flow_in.is_active is not None:
        flow.is_active = flow_in.is_active
    if flow_in.priority is not None:
        flow.priority = flow_in.priority
    if flow_in.message_priority is not None:
        flow.message_priority = flow_in.message_priority
    if flow_in.priority_over_autoreply is not None:
        flow.priority_over_autoreply = flow_in.priority_over_autoreply
    if flow_in.timeout_minutes is not None:
        flow.timeout_minutes = flow_in.timeout_minutes
        
    if flow_in.phone_ids is not None:
        _sync_phone_links(session, flow.id, flow_in.phone_ids)
        
    flow.updated_at = datetime.now()
    session.add(flow)

    if flow_in.nodes is not None and flow_in.edges is not None:
        _validate_exactly_one_start_node(flow_in.nodes)
        _validate_action_nodes(flow_in.nodes, current_user)
        
        # If user explicitly sent an empty nodes list during UPDATE, that's invalid since a real update must have ≥ 1 start node!
        if len(flow_in.nodes) == 0:
             raise HTTPException(status_code=400, detail="Flow graph cannot be empty. Must contain exactly 1 start node.")
             
        _sync_graph(session, flow.id, flow_in.nodes, flow_in.edges)

    session.commit()
    return _build_flow_response(flow, session)


@router.delete("/{flow_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(rate_limit_by_user(20, 60, "flow-delete"))])
def delete_flow(flow_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Delete a flow and its entire graph efficiently."""
    flow = session.get(Flow, flow_id)
    if not flow or flow.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Flow not found")
        
    _sync_graph(session, flow.id, [], []) # Deletes graph points using bulk SQL
    session.delete(flow)
    session.commit()
    return None


# ─────────────────────────────────────────────────────────
# Flow Sharing (Clone via Share Code)
# ─────────────────────────────────────────────────────────

import secrets

@router.post("/{flow_id}/share")
def generate_share_code(flow_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Generate a share code for this flow so other users can import a copy."""
    flow = session.get(Flow, flow_id)
    if not flow or flow.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Flow not found")
    
    if not flow.share_code:
        flow.share_code = secrets.token_urlsafe(6)  # ~8 char code
    
    session.add(flow)
    session.commit()
    session.refresh(flow)
    return {"share_code": flow.share_code}


@router.delete("/{flow_id}/share", status_code=status.HTTP_204_NO_CONTENT)
def revoke_share_code(flow_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Revoke the share code so nobody else can import this flow."""
    flow = session.get(Flow, flow_id)
    if not flow or flow.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Flow not found")
    
    flow.share_code = None
    session.add(flow)
    session.commit()


from uuid import uuid4 as _uuid4

@router.post("/import/{share_code}", response_model=FlowRead, status_code=status.HTTP_201_CREATED)
def import_flow(share_code: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Clone a shared flow into the current user's account."""
    source_flow = session.exec(select(Flow).where(Flow.share_code == share_code)).first()
    if not source_flow:
        raise HTTPException(status_code=404, detail="Invalid or expired share code")
    
    # 1. Create a new flow shell under this user
    new_flow = Flow(
        user_id=current_user.id,
        name=f"{source_flow.name} (نسخة)",
        is_active=False,
        priority=source_flow.priority,
        message_priority=source_flow.message_priority,
        priority_over_autoreply=source_flow.priority_over_autoreply,
        timeout_minutes=source_flow.timeout_minutes,
    )
    session.add(new_flow)
    session.commit()
    session.refresh(new_flow)
    
    # 2. Clone all nodes with fresh client IDs
    source_nodes = session.exec(select(FlowNode).where(FlowNode.flow_id == source_flow.id)).all()
    old_to_new_node_id = {}  # old DB id -> new DB id
    old_to_new_client_id = {}  # old client_node_id -> new client_node_id
    
    for src_node in source_nodes:
        new_client_id = str(_uuid4())
        old_to_new_client_id[src_node.client_node_id] = new_client_id
        
        new_node = FlowNode(
            flow_id=new_flow.id,
            client_node_id=new_client_id,
            node_type=src_node.node_type,
            node_data=src_node.node_data.copy() if src_node.node_data else {},
        )
        session.add(new_node)
        session.flush()
        old_to_new_node_id[src_node.id] = new_node.id
    
    # 3. Clone all edges using the new node IDs
    source_edges = session.exec(select(FlowEdge).where(FlowEdge.flow_id == source_flow.id)).all()
    for src_edge in source_edges:
        if src_edge.from_node_id in old_to_new_node_id and src_edge.to_node_id in old_to_new_node_id:
            new_edge = FlowEdge(
                flow_id=new_flow.id,
                client_edge_id=str(_uuid4()),
                from_node_id=old_to_new_node_id[src_edge.from_node_id],
                to_node_id=old_to_new_node_id[src_edge.to_node_id],
                edge_type=src_edge.edge_type,
                edge_data=src_edge.edge_data.copy() if src_edge.edge_data else {},
            )
            session.add(new_edge)
    
    session.commit()
    return _build_flow_response(new_flow, session)


# ─────────────────────────────────────────────────────────
# Flow Run Inspection & Management
# ─────────────────────────────────────────────────────────

@router.get("/runs/active")
def get_active_runs(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """List all currently running flow sessions for this user's flows."""
    # Join with Flow to ensure we only see runs belonging to the current user's flows
    runs = session.exec(
        select(FlowRun, Flow.name)
        .join(Flow, Flow.id == FlowRun.flow_id)
        .where(
            Flow.user_id == current_user.id,
            FlowRun.status == FlowRunStatus.RUNNING
        )
    ).all()
    
    return [
        {
            "id": run.id,
            "flow_id": run.flow_id,
            "flow_name": name,
            "contact_id": run.contact_id,
            "session_id": run.session_id,
            "current_node_id": run.current_node_id,
            "expires_at": run.expires_at,
            "created_at": run.created_at
        } for run, name in runs
    ]


@router.get("/runs/search/{contact_id}")
def search_run_by_contact(contact_id: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Search for any active flow run for a specific contact (phone number)."""
    # Auto-append @c.us if missing for convenience
    full_contact_id = contact_id if "@" in contact_id else f"{contact_id}@c.us"
    
    run_data = session.exec(
        select(FlowRun, Flow.name)
        .join(Flow, Flow.id == FlowRun.flow_id)
        .where(
            Flow.user_id == current_user.id,
            FlowRun.contact_id == full_contact_id,
            FlowRun.status == FlowRunStatus.RUNNING
        )
    ).first()
    
    if not run_data:
        raise HTTPException(status_code=404, detail=f"No active flow run found for {contact_id}")
        
    run, name = run_data
    return {
        "id": run.id,
        "flow_id": run.flow_id,
        "flow_name": name,
        "contact_id": run.contact_id,
        "session_id": run.session_id,
        "current_node_id": run.current_node_id,
        "expires_at": run.expires_at,
        "created_at": run.created_at
    }


@router.delete("/runs/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_run(run_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Force-terminate an active flow run."""
    run = session.exec(
        select(FlowRun)
        .join(Flow, Flow.id == FlowRun.flow_id)
        .where(FlowRun.id == run_id, Flow.user_id == current_user.id)
    ).first()
    
    if not run:
        raise HTTPException(status_code=404, detail="Flow run not found")
        
    session.delete(run)
    session.commit()
    return None


@router.delete("/runs/contact/{contact_id}", status_code=status.HTTP_204_NO_CONTENT)
def clear_contact_runs(contact_id: str, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """Clear ALL active flow runs for a specific contact (unblocks them)."""
    full_contact_id = contact_id if "@" in contact_id else f"{contact_id}@c.us"
    
    # Identify runs to delete
    runs = session.exec(
        select(FlowRun)
        .join(Flow, Flow.id == FlowRun.flow_id)
        .where(
            Flow.user_id == current_user.id,
            FlowRun.contact_id == full_contact_id
        )
    ).all()
    
    for r in runs:
        session.delete(r)
        
    session.commit()
    return None


# ─────────────────────────────────────────────────────────
# Flow Execution Engine Loop
# ─────────────────────────────────────────────────────────
async def webhook_flow_executor(
    incoming_message_id: str,
    session_id: str,
    contact_id: str,
    user_id: int,
    text: str,
    is_sandbox: bool = False
) -> tuple[bool, list[str]]:
    """
    Shows how the backend will process visual nodes continuously upon a user message webhook.
    It stops cycling when it runs out of nodes or hits a Condition requiring the *next* message.
    Returns (consumed_by_flow: bool, sandbox_responses: list[str])
    """
    session = next(get_session())
    sandbox_responses = []
    # 1. Fetch active state
    active_run = session.exec(
        select(FlowRun).where(
            FlowRun.session_id == session_id,
            FlowRun.contact_id == contact_id,
            FlowRun.status == FlowRunStatus.RUNNING
        )
    ).first()

    if active_run:
        # 1a. Duplicate message check (Waha retry protection)
        if active_run.last_processed_message_id == incoming_message_id:
            session.close()
            return True, sandbox_responses  # The Webhook fired twice, ignore it completely

        # 1b. Expiration Timer Check
        if active_run.expires_at and datetime.now() > active_run.expires_at:
            active_run.status = FlowRunStatus.EXPIRED
            session.add(active_run)
            session.commit()
            active_run = None  # Force starting from scratch
            
    current_node_id = None
    is_resuming = False
    matched_flow = None

    if not active_run:
        # 2. Check Triggers for a NEW flow
        # We need to test the incoming text against the START nodes of all ACTIVE flows (ordered by priority)
        # We only query active flows that are legally linked to the WAHA Session receiving the msg!
        # Join FlowNode to immediately fetch only their 'start' nodes natively in SQL
        active_flows = session.exec(
            select(Flow, FlowNode)
            .join(FlowPhoneLink, FlowPhoneLink.flow_id == Flow.id)
            .join(Phone, Phone.id == FlowPhoneLink.phone_id)
            .join(FlowNode, FlowNode.flow_id == Flow.id)
            .where(
                Phone.session_id == session_id,
                Flow.is_active == True,
                FlowNode.node_type == "start"
            )
            .order_by(Flow.priority.asc())
        ).all()
        
        matched_flow = None
        matched_start_node = None
        
        for flow, start_node in active_flows:
            
            # Evaluate Start Condition
            trigger_type = start_node.node_data.get("trigger_type", "any")
            trigger_value = str(start_node.node_data.get("value", "")).lower()
            msg_lower = text.lower()
            
            is_match = False
            if trigger_type == "any":
                is_match = True
            elif trigger_type == "equals" and msg_lower.strip() == trigger_value.strip():
                is_match = True
            elif trigger_type == "contains" and trigger_value.strip() in msg_lower:
                is_match = True

            if is_match:
                print(f"[Flow] 🎯 Trigger Match: {flow.name} (ID: {flow.id})")
                matched_flow = flow
                matched_start_node = start_node
                break # Found the highest priority flow that matches!
                
        if not matched_flow:
            session.close()
            return False, sandbox_responses # No flows match this new user, normal system AutoReply takes over.
            
        # 3. Create a brand new FlowRun for the matched flow!
        active_run = FlowRun(
            session_id=session_id,
            contact_id=contact_id,
            flow_id=matched_flow.id, 
            status=FlowRunStatus.RUNNING,
            current_node_id=matched_start_node.id,
            last_processed_message_id=incoming_message_id
        )
        session.add(active_run)
        # Flush to DB so active_run gets an ID
        session.commit()
        session.refresh(active_run)
        
        current_node_id = matched_start_node.id
        flow_message_priority = matched_flow.message_priority
    else:
        # 3. Resume from where they were waiting (The Condition Node ID)
        current_node_id = active_run.current_node_id
        is_resuming = True
        flow_parent = session.get(Flow, active_run.flow_id)
        flow_message_priority = flow_parent.message_priority if flow_parent else 50

    # 4. THE EXECUTION ENGINE LOOP
    # We use a loop so that consecutive Send Message nodes happen instantly in milliseconds!
    last_executed_node_type = None

    while current_node_id:
        node = session.get(FlowNode, current_node_id)
        
        if not node:
            break

        if node.node_type == "condition":
            # Very important: If we just jumped HERE from an Action/Message node *during this same loop cycle*,
            # we don't have the user's *next* message yet. So we MUST STOP and wait.
            # However, if we jumped here from a START node or ANOTHER CONDITION, we evaluate instantly!
            if last_executed_node_type == "message" and not is_resuming:
                active_run.current_node_id = current_node_id
                active_run.last_processed_message_id = incoming_message_id
                
                flow_parent = session.get(Flow, active_run.flow_id)
                if flow_parent.timeout_minutes:
                    active_run.expires_at = datetime.now() + timedelta(minutes=flow_parent.timeout_minutes)
                
                break # Exit the loop, wait for next Webhook!

            # 2. Evaluate the node against the current text
            operator = node.node_data.get("operator", "equals")
            condition_val = str(node.node_data.get("value", "")).lower()
            msg_lower = text.lower()
            
            is_true = False
            msg_clean = msg_lower.strip()
            val_clean = condition_val.strip()
            
            print(f"[Flow] ⚖️ Condition Check: '{msg_clean}' {operator} '{val_clean}'")

            if operator == "equals":
                is_true = (msg_clean == val_clean)
            elif operator == "contains":
                is_true = (val_clean in msg_clean)
            elif operator == "contains_any":
                words = [w.strip() for w in val_clean.split(",") if w.strip()]
                is_true = any(w in msg_clean for w in words)
            elif operator == "any_message":
                is_true = True
            
            print(f"[Flow]    Result: {is_true}")
            
            # 3. Find the next connected edge based on true/false
            edge_type = "true" if is_true else "false"
            next_edge = session.exec(
                select(FlowEdge).where(
                    FlowEdge.from_node_id == current_node_id,
                    FlowEdge.edge_type == edge_type
                )
            ).first()
            
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "condition"
                is_resuming = False
            else:
                current_node_id = None # End flow naturally

        elif node.node_type == "options":
            # Like Condition, if we just jumped from a Message, we must wait for the user to reply
            if last_executed_node_type == "message" and not is_resuming:
                active_run.current_node_id = current_node_id
                active_run.last_processed_message_id = incoming_message_id
                
                flow_parent = session.get(Flow, active_run.flow_id)
                if flow_parent.timeout_minutes:
                    active_run.expires_at = datetime.now() + timedelta(minutes=flow_parent.timeout_minutes)
                
                break

            msg_clean = text.lower().strip()
            options = node.node_data.get("options", [])
            print(f"[Flow] 🔀 Options Check against '{msg_clean}': {len(options)} options")
            
            matched_handle = "else"
            for opt in options:
                opt_operator = opt.get("operator", "equals")
                opt_val = str(opt.get("value", "")).lower().strip()
                
                is_true = False
                if opt_operator == "equals":
                    is_true = (msg_clean == opt_val)
                elif opt_operator == "contains":
                    # Check if text contains the val
                    is_true = (opt_val in msg_clean) if opt_val else False
                elif opt_operator == "contains_any":
                    words = [w.strip() for w in opt_val.split(",") if w.strip()]
                    is_true = any(w in msg_clean for w in words) if words else False
                elif opt_operator == "any_message":
                    is_true = True
                    
                if is_true:
                    matched_handle = opt.get("id")
                    print(f"[Flow]    Match found! Option ID: {matched_handle}")
                    break
                    
            if matched_handle == "else":
                print("[Flow]    No match. Resorting to 'else' fallback.")
                
            # Find the next connected edge for the winning Handle
            edges_out = session.exec(select(FlowEdge).where(FlowEdge.from_node_id == current_node_id)).all()
            
            next_edge = None
            for e in edges_out:
                if e.edge_data.get("_sourceHandle") == matched_handle:
                    next_edge = e
                    break
                    
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "options"
                is_resuming = False
            else:
                current_node_id = None # End flow naturally

        elif node.node_type == "start":
            # Start nodes do nothing special in the loop except instantly jump to whatever comes next!
            next_edge = session.exec(select(FlowEdge).where(
                FlowEdge.from_node_id == current_node_id
            )).first()
            
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "start"
                is_resuming = False
            else:
                current_node_id = None # Dead end flow!

        elif node.node_type == "action_ticket":
            # Action nodes do not pause! They execute logic using current memory bounds bridging right away.
            category_id = node.node_data.get("category_id")
            if category_id:
                try:
                    category_id = int(category_id)
                except ValueError:
                    category_id = None
            
            flow_parent = session.get(Flow, active_run.flow_id)
            if flow_parent:
                from app.api.tickets.routes import create_ticket_internal
                try:
                    create_ticket_internal(
                       session=session,
                       user_id=flow_parent.user_id,
                       sender_number=contact_id,
                       body=text,
                       category_id=category_id,
                       session_id=session_id
                    )
                    print(f"[Flow] 🎫 Action Node: Instant Ticket generated for {contact_id}")
                except Exception as e:
                    print(f"[Flow] ❌ Failed to create ticket in flow Action Node: {e}")
            
            # Immediately continue to the next block silently.
            next_edge = session.exec(select(FlowEdge).where(
                FlowEdge.from_node_id == current_node_id
            )).first()
            
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "action_ticket"
                is_resuming = False
            else:
                current_node_id = None
                
        elif node.node_type == "wait_message":
            if last_executed_node_type == "message" and not is_resuming:
                active_run.current_node_id = current_node_id
                active_run.last_processed_message_id = incoming_message_id
                
                flow_parent = session.get(Flow, active_run.flow_id)
                if flow_parent.timeout_minutes:
                    active_run.expires_at = datetime.now() + timedelta(minutes=flow_parent.timeout_minutes)
                
                break
                
            print(f"[Flow] ⏳ Wait Message Node fulfilled by '{text.strip()}'")
            next_edge = session.exec(select(FlowEdge).where(
                FlowEdge.from_node_id == current_node_id
            )).first()
            
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "wait_message"
                is_resuming = False
            else:
                current_node_id = None

        elif node.node_type == "action_xlsx_search":
            # Spreadsheet Search Node
            file_id = node.node_data.get("file_id")
            search_col = node.node_data.get("search_column")
            result_col = node.node_data.get("result_column")
            variable_name = node.node_data.get("variable_name", "search_result")
            
            # If search_value is empty, use the current message text
            search_val = node.node_data.get("search_value")
            if not search_val:
                search_val = text
            else:
                search_val = _resolve_variables(search_val, active_run, text)
            
            search_val_processed = str(search_val).strip().lower()
            
            found_value = None
            is_match = False
            
            if file_id and search_col and result_col:
                stored_file = session.get(StoredFile, int(file_id))
                if stored_file:
                    file_path = USER_FILES_ROOT.parent / stored_file.relative_path
                    if file_path.exists():
                        try:
                            # Use data_only to get calculated values if any
                            wb = openpyxl.load_workbook(file_path, data_only=True)
                            ws = wb.active
                            
                            # Find columns by header (case-insensitive)
                            # Assume first row contains headers
                            headers = []
                            for cell in ws[1]:
                                headers.append(str(cell.value or "").strip().lower())
                            
                            try:
                                search_idx = headers.index(str(search_col).strip().lower())
                                result_idx = headers.index(str(result_col).strip().lower())
                                
                                # Search rows (start from row 2)
                                for row in ws.iter_rows(min_row=2):
                                    cell_val = str(row[search_idx].value or "").strip().lower()
                                    if cell_val == search_val_processed:
                                        found_value = str(row[result_idx].value or "")
                                        is_match = True
                                        break
                            except ValueError:
                                print(f"[Flow] ❌ Action Node: Column not found in XLSX: '{search_col}' or '{result_col}'")
                        except Exception as e:
                            print(f"[Flow] ❌ Action Node: Error reading XLSX: {e}")
            
            if is_match:
                print(f"[Flow] 🔍 Action Node: Match found for '{search_val_processed}' -> '{found_value}'")
                # Update session memory
                updated_meta = dict(active_run.session_metadata or {})
                updated_meta[variable_name] = found_value
                active_run.session_metadata = updated_meta
                session.add(active_run)
                edge_type_to_match = "true"
            else:
                print(f"[Flow] 🔍 Action Node: No match found for '{search_val_processed}'")
                edge_type_to_match = "false"
            
            # Find next connected edge
            next_edge = session.exec(select(FlowEdge).where(
                FlowEdge.from_node_id == current_node_id,
                FlowEdge.edge_type == edge_type_to_match
            )).first()
            
            # Fallback: if no true/false edges, take any first edge
            if not next_edge:
                next_edge = session.exec(select(FlowEdge).where(
                    FlowEdge.from_node_id == current_node_id
                )).first()
            
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "action_xlsx_search"
                is_resuming = False
            else:
                current_node_id = None

        elif node.node_type == "message":
            if is_sandbox:
                sandbox_responses.append(node.node_data.get("text", ""))
            else:
                # Instantly insert into Outbox so the user receives it via Waha
                resolved_text = _resolve_variables(node.node_data.get("text", ""), active_run, text)
                outbox_id = insert_outbox(
                    session_id=session_id,
                    payload={
                        "to": contact_id,
                        "text": resolved_text
                    },
                    scheduled_at=datetime.now(timezone.utc),
                    user_id=user_id,
                    priority=flow_message_priority,
                    source_feature="flow"
                )
                print(f"[Flow] 📤 Inserted Outbox ID: {outbox_id}")

            # Find next node! (Usually connected with 'default' edge, but fallback to any)
            next_edge = session.exec(
                select(FlowEdge).where(FlowEdge.from_node_id == current_node_id)
            ).first()
            
            if next_edge:
                current_node_id = next_edge.to_node_id
                last_executed_node_type = "message"
                is_resuming = False
            else:
                current_node_id = None

        elif node.node_type == "end":
            active_run.status = FlowRunStatus.COMPLETED
            active_run.current_node_id = None
            break

        else:
            break # Handle unknown nodes

    # End of while loop processing
    if not current_node_id and active_run:
        # Flow naturally ended (no more nodes connected)
        active_run.status = FlowRunStatus.COMPLETED

    session.commit()
    
    # If the flow completed successfully or paused at a condition,
    # return True so Webhook knows the Flow consumed this message!
    result = False
    if matched_flow:
        result = matched_flow.priority_over_autoreply
    elif active_run:
        flow_parent = session.get(Flow, active_run.flow_id)
        result = flow_parent.priority_over_autoreply if flow_parent else True

    session.close()
    return result, sandbox_responses

