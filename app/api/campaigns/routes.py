from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from typing import List, Optional
import json
import re
import phonenumbers
from io import BytesIO
from sqlmodel import Session, select, update
from sqlalchemy import case, func, or_
from app.schemas.campaign import CampaignCreate, CampaignUpdate, CampaignRead, CampaignRecipientRead, CampaignResendRequest
from app.models.outbox import OutboxMessage
from app.api.deps import get_session, get_current_user, require_feature
from app.core.features import Feature
from app.core.storage import resolve_storage_path
from app.models.campaign import Campaign, CampaignRecipient
from app.models.template import Template, TemplateGroup, TemplateGroupLink
from app.models.phone import Phone, Group, PhoneGroupLink
from app.models.storage import StoredFile
from app.models.user import User
from app.models.blacklist import Blacklist
from datetime import datetime, timezone
from openpyxl import load_workbook
from app.api.rate_limit import rate_limit_by_user


router = APIRouter(dependencies=[Depends(require_feature(Feature.campaigns))])


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _extract_variables(template_body: str) -> List[str]:
    """Extract {{variable}} names from a template body."""
    return re.findall(r"\{\{(\w+)\}\}", template_body)


def _resolve_phone_ids(
    session: Session,
    phone_ids: Optional[List[int]],
    phone_group_ids: Optional[List[int]],
    user_id: int
) -> List[int]:
    """Resolve phone IDs and phone group IDs into a flat list of phone IDs."""
    resolved = set()

    # Add individual phone IDs
    if phone_ids:
        for pid in phone_ids:
            phone = session.get(Phone, pid)
            if not phone or phone.user_id != user_id:
                raise HTTPException(status_code=400, detail=f"Phone {pid} not found or not owned by you")
            if phone.status != "WORKING":
                raise HTTPException(status_code=400, detail=f"Phone '{phone.name or pid}' is not in a working state (status: {phone.status})")
            resolved.add(pid)

    # Resolve phone group IDs to individual phone IDs
    if phone_group_ids:
        for gid in phone_group_ids:
            group = session.get(Group, gid)
            if not group or group.user_id != user_id:
                raise HTTPException(status_code=400, detail=f"Phone group {gid} not found or not owned by you")
            links = session.exec(
                select(PhoneGroupLink).where(PhoneGroupLink.group_id == gid)
            ).all()
            for link in links:
                phone = session.get(Phone, link.phone_id)
                if not phone or phone.status != "WORKING":
                    phone_name = phone.name if phone else link.phone_id
                    status = phone.status if phone else "UNKNOWN"
                    raise HTTPException(status_code=400, detail=f"Phone '{phone_name}' in group '{group.name}' is not in a working state (status: {status})")
                resolved.add(link.phone_id)

    if not resolved:
        raise HTTPException(status_code=400, detail="At least one sender phone is required")

    return list(resolved)


def _get_template_bodies(
    session: Session,
    campaign_data: CampaignCreate,
    user_id: int
) -> List[str]:
    """Get template body/bodies depending on use_group flag."""
    if campaign_data.use_group:
        if not campaign_data.template_group_id:
            raise HTTPException(status_code=400, detail="template_group_id required when use_group is true")
        group = session.get(TemplateGroup, campaign_data.template_group_id)
        if not group or group.user_id != user_id:
            raise HTTPException(status_code=404, detail="Template group not found")
        links = session.exec(
            select(TemplateGroupLink).where(TemplateGroupLink.template_group_id == group.id)
        ).all()
        if not links:
            raise HTTPException(status_code=400, detail="Template group is empty")
        bodies = []
        for link in links:
            template = session.get(Template, link.template_id)
            if template:
                bodies.append(template.body)
        return bodies
    else:
        if not campaign_data.template_id:
            raise HTTPException(status_code=400, detail="template_id required when use_group is false")
        template = session.get(Template, campaign_data.template_id)
        if not template or template.user_id != user_id:
            raise HTTPException(status_code=404, detail="Template not found")
        return [template.body]


def _parse_xlsx(file_bytes: bytes, sheet_name: Optional[str] = None) -> tuple[List[str], List[dict]]:
    """Parse XLSX and return (headers, rows as list of dicts)."""
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        # Try to find first non-empty row? For now assume row 1
        raise HTTPException(status_code=400, detail="XLSX sheet is empty")

    headers = [str(h).strip() if h else "" for h in headers_raw]

    # Parse data rows
    data_rows = []
    for row in rows_iter:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = str(val) if val is not None else ""
        if any(row_dict.values()):  # skip completely empty rows
            data_rows.append(row_dict)

    wb.close()
    return headers, data_rows



def _validate_phone_for_campaign(raw: str) -> tuple[str | None, str | None]:
    raw = (raw or '').strip()
    if not raw:
        return None, None

    normalized_input = re.sub(r"[^\d+]", "", raw)
    if not normalized_input:
        return None, 'Invalid phone format'
    if not normalized_input.startswith('+'):
        normalized_input = f"+{normalized_input}"

    try:
        parsed = phonenumbers.parse(normalized_input, None)
    except phonenumbers.NumberParseException:
        return None, 'Invalid phone format'
    if not phonenumbers.is_valid_number(parsed):
        return None, 'Invalid phone number'
    normalized = f"{parsed.country_code}{parsed.national_number}"
    normalized = normalized.lstrip("+")
    return normalized, None

def _build_campaign_response(campaign: Campaign, recipient_count: int) -> dict:
    """Build a campaign response dict."""
    return {
        "id": campaign.id,
        "name": campaign.name,
        "description": campaign.description,
        "status": campaign.status,
        "template_id": campaign.template_id,
        "template_group_id": campaign.template_group_id,
        "use_group": campaign.use_group,
        "sender_phone_ids": json.loads(campaign.sender_phone_ids),
        "scheduled_at": campaign.scheduled_at,
        "created_at": campaign.created_at,
        "recipient_count": recipient_count,
    }


def _build_recipient_filters(campaign_id: int, status_filter: Optional[str], search: Optional[str]):
    filters = [CampaignRecipient.campaign_id == campaign_id]

    normalized_status = (status_filter or "all").strip().lower()
    if normalized_status and normalized_status != "all":
        filters.append(CampaignRecipient.status == normalized_status)

    normalized_search = (search or "").strip()
    if normalized_search:
        pattern = f"%{normalized_search}%"
        filters.append(
            or_(
                CampaignRecipient.phone_number.ilike(pattern),
                CampaignRecipient.rendered_message.ilike(pattern),
                CampaignRecipient.error_message.ilike(pattern),
            )
        )

    return filters


def _build_recipient_order(sort_by: Optional[str], sort_dir: str):
    direction = (sort_dir or "desc").lower()
    descending = direction == "desc"

    sort_column = {
        "phone_number": CampaignRecipient.phone_number,
        "rendered_message": CampaignRecipient.rendered_message,
        "status": CampaignRecipient.status,
        "updated_at": CampaignRecipient.updated_at,
        "error_message": CampaignRecipient.error_message,
        "sender": func.coalesce(CampaignRecipient.sent_by_session_name, CampaignRecipient.sent_by_number, ""),
    }.get(sort_by or "updated_at", CampaignRecipient.updated_at)

    return sort_column.desc() if descending else sort_column.asc()


def _get_campaign_summary(session: Session, campaign_id: int) -> dict:
    counts = session.exec(
        select(CampaignRecipient.status, func.count(CampaignRecipient.id))
        .where(CampaignRecipient.campaign_id == campaign_id)
        .group_by(CampaignRecipient.status)
    ).all()

    summary = {
        "total": 0,
        "sent": 0,
        "failed": 0,
        "pending": 0,
        "cancelled": 0,
        "skipped": 0,
    }

    for status_name, count in counts:
        summary["total"] += count
        if status_name in summary:
            summary[status_name] = count

    return summary


# ──────────────────────────────────────────────
# Campaign CRUD
# ──────────────────────────────────────────────

@router.post(
    "/",
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(rate_limit_by_user(10, 60, "campaign-create"))],
)
async def create_campaign(
    file: UploadFile | None = File(default=None),
    campaign_data: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """
    Create a campaign with an XLSX file.
    campaign_data is a JSON string of CampaignCreate fields.
    """
    # Parse campaign JSON from form
    try:
        data = CampaignCreate(**json.loads(campaign_data))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid campaign data: {str(e)}")

    if file is not None and data.stored_file_id is not None:
        raise HTTPException(status_code=400, detail="Choose either an uploaded file or a stored file")

    if file is None and data.stored_file_id is None:
        raise HTTPException(status_code=400, detail="A recipient file is required")

    if data.stored_file_id is not None:
        stored_file = session.get(StoredFile, data.stored_file_id)
        if not stored_file or stored_file.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="Stored file not found")
        if stored_file.expires_at < datetime.utcnow():
            raise HTTPException(status_code=400, detail="The selected stored file has expired")

        file_path = resolve_storage_path(stored_file.relative_path)
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Stored file content not found")

        if not stored_file.original_name.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Stored file must be .xlsx or .xls")

        file_bytes = file_path.read_bytes()
    else:
        assert file is not None
        if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
        file_bytes = await file.read()

    # Parse XLSX
    headers, data_rows = _parse_xlsx(file_bytes, data.sheet_name)

    # Validate phone column exists
    if data.phone_column not in headers:
        raise HTTPException(
            status_code=400,
            detail=f"Phone column '{data.phone_column}' not found in XLSX. Available columns: {headers}"
        )

    # Get template bodies and extract variables
    bodies = _get_template_bodies(session, data, current_user.id)
    all_variables = set()
    for body in bodies:
        all_variables.update(_extract_variables(body))

    # Validate variable mapping covers all template variables
    if all_variables:
        missing = all_variables - set(data.variable_mapping.keys())
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing mapping for template variables: {missing}. Template needs: {all_variables}"
            )

        # Validate mapped columns exist in XLSX
        for var, col in data.variable_mapping.items():
            if col not in headers:
                raise HTTPException(
                    status_code=400,
                    detail=f"Mapped column '{col}' for variable '{var}' not found in XLSX. Available: {headers}"
                )

    # Resolve sender phone IDs
    resolved_phones = _resolve_phone_ids(session, data.phone_ids, data.phone_group_ids, current_user.id)

    invalid_numbers = []
    normalized_rows = []
    for row in data_rows:
        raw_phone = str(row.get(data.phone_column, '')).strip()
        normalized_phone, reason = _validate_phone_for_campaign(raw_phone)
        if reason:
            invalid_numbers.append({"value": raw_phone, "reason": reason})
            continue
        if not normalized_phone:
            continue
        normalized_rows.append((row, normalized_phone))
    blacklisted_numbers = set(
        session.exec(select(Blacklist.phone_number).where(Blacklist.user_id == current_user.id)).all()
    )

    deliverable_rows = [
        (row, phone_number)
        for row, phone_number in normalized_rows
        if phone_number not in blacklisted_numbers
    ]

    if not deliverable_rows:
        raise HTTPException(
            status_code=400,
            detail="No valid recipients found after filtering invalid and blacklisted numbers",
        )

    # Create campaign
    campaign = Campaign(
        name=data.name,
        description=data.description,
        template_id=data.template_id if not data.use_group else None,
        template_group_id=data.template_group_id if data.use_group else None,
        use_group=data.use_group,
        sender_phone_ids=json.dumps(resolved_phones),
        phone_column=data.phone_column,
        variable_mapping=json.dumps(data.variable_mapping),
        scheduled_at=data.scheduled_at,
        user_id=current_user.id,
    )
    session.add(campaign)
    session.commit()
    session.refresh(campaign)

    # Create recipients from XLSX rows
    recipients = []
    outbox_messages = []
    
    # Fetch resolved phones to get their session_ids
    phones = session.exec(select(Phone).where(Phone.id.in_(resolved_phones))).all()
    if not phones:
        raise HTTPException(status_code=400, detail="No valid sender phones found")
    
    phone_map = {p.id: p.session_id for p in phones}
    # Ensure ordered list for round-robin
    sender_session_ids = [phone_map[pid] for pid in resolved_phones if pid in phone_map]

    # Calculate scheduled_time
    # now = datetime.now(timezone.utc) # Not needed for delay anymore
    if data.scheduled_at.tzinfo is None:
        # If naive, assume it's meant to be UTC
        scheduled_at_aware = data.scheduled_at.replace(tzinfo=timezone.utc)
    else:
        scheduled_at_aware = data.scheduled_at
        
    # Delay calculation removed as per user request

    from app.api.outbox.routes import bulk_insert_outbox

    for i, item in enumerate(deliverable_rows):
        row, phone_number = item

        # Render message
        template_body = bodies[i % len(bodies)]
        rendered_msg = template_body
        for var, col in data.variable_mapping.items():
            val = row.get(col, "")
            rendered_msg = rendered_msg.replace(f"{{{{{var}}}}}", str(val))

        recipient = CampaignRecipient(
            phone_number=phone_number,
            row_data=json.dumps(row),
            campaign_id=campaign.id,
            rendered_message=rendered_msg,
        )
        recipients.append(recipient)
        session.add(recipient)
        
        # Assign sender session_id (round-robin)
        assigned_idx = i % len(sender_session_ids)
        session_id = sender_session_ids[assigned_idx]
        
        # Rest of the chosen numbers become backups for this specific message in case it fails
        fallback_ids = [sid for idx, sid in enumerate(sender_session_ids) if idx != assigned_idx]
        
        outbox_messages.append({
            "session_id": session_id,
            "fallback_session_ids": fallback_ids,
            "payload": {
                "to": phone_number,
                "text": rendered_msg,
            },
            "scheduled_at": scheduled_at_aware,
            "user_id": current_user.id,
            "priority": 100,
            "source_feature": Feature.campaigns.value,
            "campaign_id": campaign.id,
            "recipient_index": len(recipients) - 1,  # track which recipient this belongs to
        })

    # Flush to get recipient IDs
    session.flush()

    # Now inject campaign_recipient_id into each outbox message payload
    for msg in outbox_messages:
        idx = msg.pop("recipient_index")
        recipient = recipients[idx]
        msg["payload"]["campaign_recipient_id"] = recipient.id

    session.commit()
    
    # Bulk insert into outbox
    if outbox_messages:
        bulk_insert_outbox(outbox_messages)
        
    # Update campaign status to pending
    campaign.status = "scheduled"
    session.add(campaign)
    session.commit()
    session.refresh(campaign)

    return {"success": True}


@router.get("/", response_model=List[CampaignRead])
def get_campaigns(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """List all campaigns for the current user."""
    campaigns = session.exec(
        select(Campaign).where(Campaign.user_id == current_user.id)
    ).all()

    campaign_ids = [campaign.id for campaign in campaigns if campaign.id is not None]
    counts_by_campaign = {}
    pending_by_campaign = {}
    if campaign_ids:
        stats_rows = session.exec(
            select(
                CampaignRecipient.campaign_id,
                func.count(CampaignRecipient.id),
                func.sum(case((CampaignRecipient.status == "pending", 1), else_=0)),
            )
            .where(CampaignRecipient.campaign_id.in_(campaign_ids))
            .group_by(CampaignRecipient.campaign_id)
        ).all()
        counts_by_campaign = {campaign_id: total for campaign_id, total, _pending in stats_rows}
        pending_by_campaign = {campaign_id: pending or 0 for campaign_id, _total, pending in stats_rows}

    result = []
    dirty = False
    for c in campaigns:
        count = counts_by_campaign.get(c.id, 0)
        
        # Determine if finished
        if c.status == "scheduled":
            pending_count = pending_by_campaign.get(c.id, 0)
            if count > 0 and pending_count == 0:
                c.status = "finished"
                session.add(c)
                dirty = True
                
        result.append(_build_campaign_response(c, count))
        
    if dirty:
        session.commit()
        
    return result


@router.get("/{campaign_id}")
def get_campaign(
    campaign_id: int,
    page: int = 1,
    page_size: int = 100,
    status_filter: str = "all",
    search: Optional[str] = None,
    sort_by: str = "updated_at",
    sort_dir: str = "desc",
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Get a single campaign with summary data and a paginated recipient report."""
    campaign = session.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)

    summary = _get_campaign_summary(session, campaign.id)
    total = summary["total"]
    pending = summary["pending"]

    if campaign.status == "scheduled" and total > 0 and pending == 0:
        campaign.status = "finished"
        session.add(campaign)
        session.commit()

    filters = _build_recipient_filters(campaign.id, status_filter, search)
    filtered_total = session.exec(
        select(func.count(CampaignRecipient.id)).where(*filters)
    ).one()
    total_pages = max((filtered_total + page_size - 1) // page_size, 1)
    page = min(page, total_pages)

    recipients = session.exec(
        select(CampaignRecipient)
        .where(*filters)
        .order_by(_build_recipient_order(sort_by, sort_dir), CampaignRecipient.id.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()

    return {
        "campaign": _build_campaign_response(campaign, total),
        "summary": summary,
        "recipients": [
            {
                "id": r.id,
                "phone_number": r.phone_number,
                "rendered_message": r.rendered_message,
                "status": r.status,
                "error_message": r.error_message,
                "sent_by_session_name": r.sent_by_session_name,
                "sent_by_number": r.sent_by_number,
                "updated_at": r.updated_at,
            }
            for r in recipients
        ],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "filtered_total": filtered_total,
            "total_pages": total_pages,
        },
    }


@router.put("/{campaign_id}", dependencies=[Depends(rate_limit_by_user(20, 60, "campaign-update"))])
def update_campaign(
    campaign_id: int,
    campaign_in: CampaignUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Update campaign name, description, status, or scheduled_at."""
    campaign = session.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if campaign_in.name is not None:
        campaign.name = campaign_in.name
    if campaign_in.description is not None:
        campaign.description = campaign_in.description
    if campaign_in.status is not None:
        campaign.status = campaign_in.status
    if campaign_in.scheduled_at is not None:
        if campaign_in.scheduled_at.tzinfo is None:
            campaign.scheduled_at = campaign_in.scheduled_at.replace(tzinfo=timezone.utc)
        else:
            campaign.scheduled_at = campaign_in.scheduled_at

    session.add(campaign)
    session.commit()
    session.refresh(campaign)

    return {"success": True}


@router.delete(
    "/{campaign_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(rate_limit_by_user(5, 60, "campaign-delete"))],
)
def delete_campaign(
    campaign_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Delete a campaign and all its recipients."""
    campaign = session.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if campaign.status == "scheduled":
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete an active campaign. Please cancel it first."
        )

    # Delete recipients first
    recipients = session.exec(
        select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign.id)
    ).all()
    for r in recipients:
        session.delete(r)


    session.delete(campaign)
    session.commit()
    return None


@router.post("/{campaign_id}/cancel", dependencies=[Depends(rate_limit_by_user(5, 60, "campaign-cancel"))])
def cancel_campaign(
    campaign_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Cancel an active campaign, stopping all further queued messages."""
    campaign = session.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if campaign.status in ["finished", "cancelled"]:
        raise HTTPException(status_code=400, detail="Campaign is already finished or cancelled")

    # Update DB status
    campaign.status = "cancelled"
    session.add(campaign)
    
    # Cancel in outbox directly (both pending and already queued)
    session.exec(
        update(OutboxMessage)
        .where(OutboxMessage.campaign_id == campaign.id)
        .where(OutboxMessage.status.in_(["pending", "queued"]))
        .values(status="cancelled")
    )

    # Cancel in CampaignRecipient (for UI reporting sync)
    session.exec(
        update(CampaignRecipient)
        .where(CampaignRecipient.campaign_id == campaign.id)
        .where(CampaignRecipient.status == "pending")
        .values(status="cancelled", error_message="Campaign cancelled")
    )
    
    session.commit()

    return {"success": True}


@router.post(
    "/{campaign_id}/resend",
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(rate_limit_by_user(5, 60, "campaign-resend"))],
)
def resend_campaign_recipients(
    campaign_id: int,
    resend_data: CampaignResendRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Resend selected recipients as a new campaign."""
    # Validate original campaign
    original = session.get(Campaign, campaign_id)
    if not original:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if original.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get selected recipients (must belong to this campaign)
    selected = session.exec(
        select(CampaignRecipient).where(
            CampaignRecipient.id.in_(resend_data.recipient_ids),
            CampaignRecipient.campaign_id == campaign_id,
        )
    ).all()
    if not selected:
        raise HTTPException(status_code=400, detail="No valid recipients found")

    # Resolve sender phones (validates WORKING status)
    resolved_phones = _resolve_phone_ids(session, resend_data.phone_ids, resend_data.phone_group_ids, current_user.id)

    # Filter out blacklisted numbers
    blacklisted = set(
        session.exec(select(Blacklist.phone_number).where(Blacklist.user_id == current_user.id)).all()
    )
    deliverable = [r for r in selected if r.phone_number not in blacklisted]
    if not deliverable:
        raise HTTPException(status_code=400, detail="All selected recipients are blacklisted")

    # Schedule time
    scheduled_at = resend_data.scheduled_at
    if scheduled_at.tzinfo is None:
        scheduled_at = scheduled_at.replace(tzinfo=timezone.utc)

    # Create new campaign (copy immutable fields from original)
    new_campaign = Campaign(
        name=f"إعادة إرسال - {original.name}",
        description=original.description,
        template_id=original.template_id,
        template_group_id=original.template_group_id,
        use_group=original.use_group,
        sender_phone_ids=json.dumps(resolved_phones),
        phone_column=original.phone_column,
        variable_mapping=original.variable_mapping,
        scheduled_at=scheduled_at,
        user_id=current_user.id,
    )
    session.add(new_campaign)
    session.commit()
    session.refresh(new_campaign)

    # Get sender session_ids for round-robin
    phones = session.exec(select(Phone).where(Phone.id.in_(resolved_phones))).all()
    phone_map = {p.id: p.session_id for p in phones}
    sender_session_ids = [phone_map[pid] for pid in resolved_phones if pid in phone_map]

    # Create recipients and outbox messages
    recipients = []
    outbox_messages = []
    for i, orig_r in enumerate(deliverable):
        recipient = CampaignRecipient(
            phone_number=orig_r.phone_number,
            row_data=orig_r.row_data,
            campaign_id=new_campaign.id,
            rendered_message=orig_r.rendered_message,
        )
        recipients.append(recipient)
        session.add(recipient)

        assigned_idx = i % len(sender_session_ids)
        session_id = sender_session_ids[assigned_idx]
        fallback_ids = [sid for idx, sid in enumerate(sender_session_ids) if idx != assigned_idx]

        outbox_messages.append({
            "session_id": session_id,
            "fallback_session_ids": fallback_ids,
            "payload": {
                "to": orig_r.phone_number,
                "text": orig_r.rendered_message,
            },
            "scheduled_at": scheduled_at,
            "user_id": current_user.id,
            "priority": 100,
            "source_feature": Feature.campaigns.value,
            "campaign_id": new_campaign.id,
            "recipient_index": len(recipients) - 1,
        })

    session.flush()

    for msg in outbox_messages:
        idx = msg.pop("recipient_index")
        recipient = recipients[idx]
        msg["payload"]["campaign_recipient_id"] = recipient.id

    session.commit()

    from app.api.outbox.routes import bulk_insert_outbox
    if outbox_messages:
        bulk_insert_outbox(outbox_messages)

    new_campaign.status = "scheduled"
    session.add(new_campaign)
    session.commit()
    session.refresh(new_campaign)

    return {"success": True}
